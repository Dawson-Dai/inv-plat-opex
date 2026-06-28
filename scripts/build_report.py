"""
Build an Excel maturity report from a sorted Cortex maturity CSV.
Produces: Tribe Overview sheet + one sheet per squad.

NOTE: Input must be the sorted/filtered CSV (output of sort_csv.py).
Passing the raw Cortex export directly may include excluded squads.

Styling matches the 2026-06-20 reference report exactly.
"""
import csv
import io
import re
import sys
import zipfile
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ── Colours (exact RGB from 2026-06-20 reference) ───────────────────────────
def _fill(hex6):
    return PatternFill("solid", fgColor="00" + hex6)

FILL_TITLE    = _fill("0D1B2A")   # very dark navy – title row
FILL_SC_GRP   = _fill("2E5F9B")   # mid blue      – scorecard group headers
FILL_SUB_HDR  = _fill("4A86C8")   # light blue    – sub-headers / column headers
FILL_WHITE    = _fill("FFFFFF")   # white          – alternating row colour B
FILL_SECTION  = _fill("1F3A5F")   # dark blue     – section headings in squad sheets
FILL_ALT      = _fill("EBF2FB")   # very light    – alternating data rows
FILL_TOTAL    = _fill("BDD7EE")   # pale blue     – total rows

# ── Fonts ────────────────────────────────────────────────────────────────────
F_WHITE_BOLD_14 = Font(bold=True, size=14, color="00FFFFFF")   # title
F_WHITE_BOLD_13 = Font(bold=True, size=13, color="00FFFFFF")   # squad title
F_WHITE_BOLD_11 = Font(bold=True, size=11, color="00FFFFFF")   # section headings
F_WHITE_BOLD_9  = Font(bold=True, size=9,  color="00FFFFFF")   # column headers / KPI hdrs
F_DARK_10       = Font(size=10, color="001F2D3D")              # normal data
F_DARK_10_BOLD  = Font(bold=True, size=10, color="001F2D3D")   # squad name in tribe matrix
F_RED_10        = Font(size=10, color="00C0392B")              # entities col in tribe matrix
F_KPI_VALUE     = Font(bold=True, size=16, color="00C0392B")   # KPI value row
F_TOTAL_DARK    = Font(bold=True, size=10, color="001F2D3D")   # total row
F_ITALIC_9      = Font(italic=True, size=9)                    # footnote

# ── Border ───────────────────────────────────────────────────────────────────
THIN   = Side(style="thin",   color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Scorecard abbreviations (matches reference) ───────────────────────────────
SC_ABBREV = {
    "Anyone Can Contribute Safely":        "Contribute Safely",
    "Changes Ship Reliably":               "Ship Reliably",
    "Customer Data Stays Protected":       "Data Protected",
    "Data, AI, and ML Governed":           "AI/ML Governed",
    "Production Observable and Resilient": "Observable",
    "Technology Stays Current":            "Tech Current",
}


def _abbrev(sc: str) -> str:
    return SC_ABBREV.get(sc, sc)


def _apply(cell, value=None, font=None, fill=None, align=None, border=True):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if align:
        cell.alignment = align
    return cell


def _load(sorted_csv: str):
    with open(sorted_csv, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_report(sorted_csv: str, output_xlsx: str) -> dict:
    rows = _load(sorted_csv)
    wb = Workbook()

    # Build data tree: squad → scorecard → rule → {entities, rows, last_eval}
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"entities": set(), "rows": 0, "last_eval": {}}
    )))
    for r in rows:
        d = data[r["Squad"]][r["Scorecard"]][r["Rule"]]
        d["entities"].add(r["Entity"])
        d["rows"] += 1
        d["last_eval"][r["Entity"]] = r.get("Last Evaluated", "")

    squads = sorted(data.keys())
    all_scorecards = sorted({sc for sq in data.values() for sc in sq})

    # Date from filename
    m = re.search(r"\d{4}-\d{2}-\d{2}", sorted_csv)
    report_date = m.group() if m else "unknown"

    # ── Tribe Overview sheet ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tribe Overview"
    # No freeze panes on Tribe Overview

    # Column layout:
    # A=Squad(38), B=Entities(10), then per-SC: Rules(8)/Entities(9)/Rows(8) × 6 SCs, Grand Total same
    COL_SQUAD    = 1   # A
    COL_ENTITIES = 2   # B
    SC_COLS = {}
    col = 3
    for sc in all_scorecards:
        SC_COLS[sc] = col
        col += 3
    GT_COL = col       # Grand Total starts here
    LAST_COL = GT_COL + 2

    # Fixed column widths (match reference exactly)
    ws.column_dimensions["A"].width = 38.0
    ws.column_dimensions["B"].width = 10.0
    num_col = 3
    for sc in all_scorecards:
        ws.column_dimensions[get_column_letter(num_col)].width = 8.0    # Rules
        ws.column_dimensions[get_column_letter(num_col+1)].width = 9.0  # Entities
        ws.column_dimensions[get_column_letter(num_col+2)].width = 8.0  # Rows
        num_col += 3
    ws.column_dimensions[get_column_letter(GT_COL)].width   = 9.0
    ws.column_dimensions[get_column_letter(GT_COL+1)].width = 8.0
    ws.column_dimensions[get_column_letter(GT_COL+2)].width = 8.0

    # Fixed row heights (match reference)
    ws.row_dimensions[1].height = 34.0
    ws.row_dimensions[2].height = 36.0
    ws.row_dimensions[3].height = 26.0
    for r in range(4, 4 + len(squads)):
        ws.row_dimensions[r].height = 20.0
    ws.row_dimensions[4 + len(squads)].height = 24.0  # TOTAL row

    # Row 1: full-width title
    c = ws.cell(1, 1,
        f"Inventory Platform  –  Baseline Maturity Failures  |  Tribe Overview  |  {report_date}")
    _apply(c, font=F_WHITE_BOLD_14, fill=FILL_TITLE, border=False,
           align=Alignment(horizontal="center", vertical="center"))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)

    # Row 2: A2:B2 empty merged, then scorecard group labels, then Grand Total
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    for sc in all_scorecards:
        c = ws.cell(2, SC_COLS[sc])
        ws.merge_cells(start_row=2, start_column=SC_COLS[sc], end_row=2, end_column=SC_COLS[sc]+2)
        _apply(c, value=_abbrev(sc), font=F_WHITE_BOLD_9, fill=FILL_SC_GRP, border=False,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    c = ws.cell(2, GT_COL)
    ws.merge_cells(start_row=2, start_column=GT_COL, end_row=2, end_column=LAST_COL)
    _apply(c, value="Grand Total", font=F_WHITE_BOLD_9, fill=FILL_TITLE, border=False,
           align=Alignment(horizontal="center", vertical="center"))

    # Row 3: sub-headers
    def _sh(r, col, val, fill=FILL_SUB_HDR):
        _apply(ws.cell(r, col), value=val, font=F_WHITE_BOLD_9, fill=fill,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True))

    _sh(3, COL_SQUAD,    "Squad")
    _sh(3, COL_ENTITIES, "Entities")
    for sc in all_scorecards:
        c0 = SC_COLS[sc]
        _sh(3, c0,   "Rules")
        _sh(3, c0+1, "Entities")
        _sh(3, c0+2, "Rows")
    _sh(3, GT_COL,   "Rules")
    _sh(3, GT_COL+1, "Entities")
    _sh(3, GT_COL+2, "Rows")

    # Data rows
    tribe_entities = set()
    tribe_rows = 0
    tribe_sc_agg = defaultdict(lambda: {"rules": set(), "entities": set(), "rows": 0})
    tribe_all_rules = set()

    for ri, squad in enumerate(squads, start=4):
        alt_fill = FILL_ALT if (ri - 4) % 2 == 0 else FILL_WHITE  # ABAB by row
        sq = data[squad]
        sq_ent = {e for sc in sq.values() for rd in sc.values() for e in rd["entities"]}
        sq_rows = sum(rd["rows"] for sc in sq.values() for rd in sc.values())
        sq_rules_count = sum(len(sc) for sc in sq.values())

        def _dc(col, val, font=F_DARK_10, align_h="center"):
            _apply(ws.cell(ri, col), value=val, font=font, fill=alt_fill,
                   align=Alignment(horizontal=align_h, vertical="center"))

        _dc(COL_SQUAD,    squad,        font=F_DARK_10_BOLD, align_h="left")
        _dc(COL_ENTITIES, len(sq_ent),  font=F_RED_10)

        for sc in all_scorecards:
            sc_rules = sq.get(sc, {})
            ru = len(sc_rules)
            en = len({e for rd in sc_rules.values() for e in rd["entities"]})
            ro = sum(rd["rows"] for rd in sc_rules.values())
            c0 = SC_COLS[sc]
            _dc(c0,   ru, font=F_DARK_10)
            _dc(c0+1, en, font=F_DARK_10)
            _dc(c0+2, ro, font=F_DARK_10)
            if ru:
                tribe_sc_agg[sc]["rules"].update(sc_rules.keys())
                tribe_sc_agg[sc]["entities"].update(
                    e for rd in sc_rules.values() for e in rd["entities"])
                tribe_sc_agg[sc]["rows"] += ro

        _dc(GT_COL,   sq_rules_count, font=F_DARK_10)
        _dc(GT_COL+1, len(sq_ent),    font=F_DARK_10)
        _dc(GT_COL+2, sq_rows,        font=F_DARK_10)

        tribe_entities.update(sq_ent)
        tribe_rows += sq_rows
        tribe_all_rules.update(r for sc in sq.values() for r in sc)

    # TOTAL row
    total_ri = 4 + len(squads)
    def _tc(col, val):
        _apply(ws.cell(total_ri, col), value=val, font=F_TOTAL_DARK, fill=FILL_TOTAL,
               align=Alignment(horizontal="center", vertical="center"))
    ws.cell(total_ri, COL_SQUAD).value = "TOTAL"
    ws.cell(total_ri, COL_SQUAD).font  = F_TOTAL_DARK
    ws.cell(total_ri, COL_SQUAD).fill  = FILL_TOTAL
    ws.cell(total_ri, COL_SQUAD).border = BORDER
    ws.cell(total_ri, COL_SQUAD).alignment = Alignment(horizontal="left", vertical="center")
    _tc(COL_ENTITIES, len(tribe_entities))
    for sc in all_scorecards:
        agg = tribe_sc_agg[sc]
        c0 = SC_COLS[sc]
        _tc(c0,   len(agg["rules"]))
        _tc(c0+1, len(agg["entities"]))
        _tc(c0+2, agg["rows"])
    _tc(GT_COL,   len(tribe_all_rules))
    _tc(GT_COL+1, len(tribe_entities))
    _tc(GT_COL+2, tribe_rows)

    # Footnote (row total+2, merged full width)
    fn_row = total_ri + 2
    ws.row_dimensions[fn_row].height = 18.0
    fn_c = ws.cell(fn_row, 1,
        "All entries are Baseline-level Fails.  "
        "Rules = distinct failing rules;  "
        "Entities = distinct services/components;  "
        "Rows = total failing rows.")
    fn_c.font = F_ITALIC_9
    ws.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=LAST_COL)

    # Tribe Scorecard Summary (starts fn_row+2)
    tss_row = fn_row + 2
    ws.row_dimensions[tss_row].height = 26.0
    tss_hdr_cell = ws.cell(tss_row, 1, "Tribe Scorecard Summary")
    tss_hdr_cell.font = F_WHITE_BOLD_11
    tss_hdr_cell.fill = FILL_TITLE
    tss_hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=tss_row, start_column=1, end_row=tss_row, end_column=6)

    tss_col_hdrs = ["Scorecard", "Failing Rules", "Affected Entities",
                    "% of Tribe Entities", "Total Rows", "% of Tribe Rows"]
    ws.row_dimensions[tss_row+1].height = 28.0
    for ci, h in enumerate(tss_col_hdrs, 1):
        _apply(ws.cell(tss_row+1, ci), value=h, font=F_WHITE_BOLD_9, fill=FILL_SUB_HDR,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True))

    for i, sc in enumerate(all_scorecards):
        agg = tribe_sc_agg[sc]
        ri2 = tss_row + 2 + i
        ws.row_dimensions[ri2].height = 18.0
        pct_en = f"{len(agg['entities'])/len(tribe_entities)*100:.1f}%" if tribe_entities else "0.0%"
        pct_ro = f"{agg['rows']/tribe_rows*100:.1f}%" if tribe_rows else "0.0%"
        tss_fill = FILL_ALT if i % 2 == 0 else FILL_WHITE
        _apply(ws.cell(ri2, 1), value=sc, font=F_DARK_10, fill=tss_fill,
               align=Alignment(horizontal="left", vertical="center"))
        for ci, v in enumerate([len(agg["rules"]), len(agg["entities"]), pct_en,
                                 agg["rows"], pct_ro], start=2):
            _apply(ws.cell(ri2, ci), value=v, font=F_DARK_10, fill=tss_fill,
                   align=Alignment(horizontal="center", vertical="center"))

    # Grand total for summary
    gt_ri = tss_row + 2 + len(all_scorecards)
    ws.row_dimensions[gt_ri].height = 22.0
    _apply(ws.cell(gt_ri, 1), value="Grand Total", font=F_TOTAL_DARK, fill=FILL_TOTAL,
           align=Alignment(horizontal="left", vertical="center"))
    for ci, v in enumerate([len(tribe_all_rules), len(tribe_entities), "100%",
                             tribe_rows, "100%"], start=2):
        _apply(ws.cell(gt_ri, ci), value=v, font=F_TOTAL_DARK, fill=FILL_TOTAL,
               align=Alignment(horizontal="center", vertical="center"))

    # Bar chart — horizontal, squad names on left, placed below Tribe Scorecard Summary.
    # openpyxl always serialises category refs as numRef and titles as inline <v>.
    # Neither shows squad name labels in Excel. We patch the XML after saving via
    # _patch_chart_xml() to use strRef exactly as the 2026-06-20 reference does.
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = "Failing Rows & Entities with Baseline Failures by Squad"
    # Match reference exactly: cx=10800000, cy=7920000 EMU = 30cm × 22cm
    chart.width  = 30.0
    chart.height = 22.0

    data_last_row = 3 + len(squads)
    ents_ref = Reference(ws, min_col=COL_ENTITIES, min_row=4, max_row=data_last_row)
    rows_ref = Reference(ws, min_col=GT_COL+2,     min_row=4, max_row=data_last_row)
    cats_ref = Reference(ws, min_col=COL_SQUAD,    min_row=4, max_row=data_last_row)

    chart.add_data(ents_ref, titles_from_data=False)
    chart.add_data(rows_ref, titles_from_data=False)
    chart.series[0].title = SeriesLabel(v="Entities")
    chart.series[1].title = SeriesLabel(v="Rows")
    chart.set_categories(cats_ref)

    # Place chart 2 rows below Grand Total of Tribe Scorecard Summary
    chart_anchor_row = gt_ri + 2
    ws.add_chart(chart, f"A{chart_anchor_row}")

    # Store patching info for post-save XML fix
    _chart_cat_range  = f"'Tribe Overview'!$A$4:$A${data_last_row}"
    _chart_title_ents = f"'Tribe Overview'!B3"
    _chart_title_rows = f"'Tribe Overview'!{get_column_letter(GT_COL+2)}3"

    # ── Per-squad sheets ─────────────────────────────────────────────────────
    for squad in squads:
        safe = (squad[:28]
                .replace("/", "-").replace("\\", "-")
                .replace("*","").replace("?","").replace("[","").replace("]","").replace(":",""))
        ws2 = wb.create_sheet(title=safe)
        sq = data[squad]

        sq_ent = {e for sc in sq.values() for rd in sc.values() for e in rd["entities"]}
        sq_rows_total = sum(rd["rows"] for sc in sq.values() for rd in sc.values())
        sq_rules_total = sum(len(sc) for sc in sq.values())
        sq_sc_count = len(sq)

        # Column layout (side-by-side):
        #   A–F  (1–6)  : Scorecard Summary
        #   G    (7)    : gap
        #   H–L  (8–12) : Rule Breakdown
        #   M    (13)   : gap
        #   N–Q  (14–17): Entity Detail
        SC_COL  = 1   # Scorecard Summary starts col A
        RB_COL  = 8   # Rule Breakdown starts col H
        ED_COL  = 14  # Entity Detail starts col N

        col_widths = {
            "A": 24.0, "B": 22.0, "C": 20.0, "D": 14.0, "E": 12.0, "F": 14.0,
            "G": 3.0,
            "H": 18.0, "I": 28.0, "J": 12.0, "K": 10.0, "L": 40.0,
            "M": 3.0,
            "N": 24.0, "O": 18.0, "P": 50.0, "Q": 20.0,
        }
        for col_letter, w in col_widths.items():
            ws2.column_dimensions[col_letter].width = w

        ws2.freeze_panes = None

        # Row 1: title (merged A:Q)
        ws2.row_dimensions[1].height = 30.0
        r1 = ws2.cell(1, 1, f"Squad: {squad}  –  Baseline Maturity Failures  |  {report_date}")
        _apply(r1, font=F_WHITE_BOLD_13, fill=FILL_TITLE, border=False,
               align=Alignment(horizontal="center", vertical="center"))
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

        # Row 2: KPI headers
        ws2.row_dimensions[2].height = 20.0
        kpi_hdrs = ["Entities", "Scorecards", "Failing Rules", "Total Rows"]
        for ci, h in enumerate(kpi_hdrs, 1):
            _apply(ws2.cell(2, ci), value=h, font=F_WHITE_BOLD_9, fill=FILL_SC_GRP,
                   align=Alignment(horizontal="center", vertical="center"))

        # Row 3: KPI values (large red numbers)
        ws2.row_dimensions[3].height = 28.0
        kpi_vals = [len(sq_ent), sq_sc_count, sq_rules_total, sq_rows_total]
        for ci, v in enumerate(kpi_vals, 1):
            _apply(ws2.cell(3, ci), value=v, font=F_KPI_VALUE, fill=FILL_ALT,
                   align=Alignment(horizontal="center", vertical="center"))

        # ── Row 5: all three section headings on the same row ────────────────
        ws2.row_dimensions[5].height = 22.0

        # Scorecard Summary heading (A:F)
        sh = ws2.cell(5, SC_COL, "Scorecard Summary")
        _apply(sh, font=F_WHITE_BOLD_11, fill=FILL_SECTION, border=False,
               align=Alignment(horizontal="left", vertical="center"))
        ws2.merge_cells(start_row=5, start_column=SC_COL, end_row=5, end_column=SC_COL + 5)

        # Rule Breakdown heading (H:L)
        rb_sh = ws2.cell(5, RB_COL,
            "Rule Breakdown  –  distinct rules failing and how many entities are affected")
        _apply(rb_sh, font=F_WHITE_BOLD_11, fill=FILL_SECTION, border=False,
               align=Alignment(horizontal="left", vertical="center"))
        ws2.merge_cells(start_row=5, start_column=RB_COL, end_row=5, end_column=RB_COL + 4)

        # Entity Detail heading (N:Q)
        ed_sh = ws2.cell(5, ED_COL, "Entity Detail  –  every failing rule per entity")
        _apply(ed_sh, font=F_WHITE_BOLD_11, fill=FILL_SECTION, border=False,
               align=Alignment(horizontal="left", vertical="center"))
        ws2.merge_cells(start_row=5, start_column=ED_COL, end_row=5, end_column=ED_COL + 3)

        # ── Row 6: all three header rows ─────────────────────────────────────
        ws2.row_dimensions[6].height = 28.0

        # Scorecard Summary headers
        sc_hdrs = ["Scorecard", "Failing Rules", "Affected Entities",
                   "% of Squad Entities", "Total Rows", "% of Squad Rows"]
        for ci, h in enumerate(sc_hdrs, SC_COL):
            _apply(ws2.cell(6, ci), value=h, font=F_WHITE_BOLD_9, fill=FILL_SUB_HDR,
                   align=Alignment(horizontal="center", vertical="center", wrap_text=True))

        # Rule Breakdown headers
        rb_hdrs = ["Scorecard", "Rule", "Failing Entities", "% of Squad", "Affected Entities (names)"]
        for ci, h in enumerate(rb_hdrs, RB_COL):
            _apply(ws2.cell(6, ci), value=h, font=F_WHITE_BOLD_9, fill=FILL_SUB_HDR,
                   align=Alignment(horizontal="center", vertical="center", wrap_text=True))

        # Entity Detail headers
        ed_hdrs = ["Entity", "Scorecard", "Rule", "Last Evaluated"]
        for ci, h in enumerate(ed_hdrs, ED_COL):
            _apply(ws2.cell(6, ci), value=h, font=F_WHITE_BOLD_9, fill=FILL_SUB_HDR,
                   align=Alignment(horizontal="center", vertical="center", wrap_text=True))

        # ── Data rows (all three tables in parallel, starting row 7) ─────────
        # Each table has independent ABAB alternating colours (FILL_ALT / FILL_WHITE).
        # Rule Breakdown: scorecard cell merged across its rules.
        # Entity Detail: entity cell merged across its rows.

        # ── Scorecard Summary rows ────────────────────────────────────────────
        sc_rows = []
        for sc in all_scorecards:
            rules = sq.get(sc, {})
            ru = len(rules)
            en = len({e for rd in rules.values() for e in rd["entities"]})
            ro = sum(rd["rows"] for rd in rules.values())
            pct_en = f"{en/len(sq_ent)*100:.1f}%" if sq_ent else "0.0%"
            pct_ro = f"{ro/sq_rows_total*100:.1f}%" if sq_rows_total else "0.0%"
            sc_rows.append((_abbrev(sc), ru, en, pct_en, ro, pct_ro, False))
        sc_rows.append(("Total", sq_rules_total, len(sq_ent), "–", sq_rows_total, "100%", True))

        # ── Rule Breakdown rows: (sc_abbrev, rule, en, pct, names, en_count, sc_group_idx)
        rb_rows = []
        rb_sc_groups = []  # list of (start_idx, length) for merge tracking
        sc_group_idx = -1
        cur_sc = None
        for sc in all_scorecards:
            rules = sq.get(sc, {})
            if not rules:
                continue
            abbr = _abbrev(sc)
            group_start = len(rb_rows)
            for rule, rd in sorted(rules.items(), key=lambda x: -len(x[1]["entities"])):
                en = len(rd["entities"])
                pct = f"{en/len(sq_ent)*100:.1f}%" if sq_ent else "0.0%"
                names = "\n".join(sorted(rd["entities"]))
                rb_rows.append((abbr, rule, en, pct, names, en))
            group_len = len(rb_rows) - group_start
            rb_sc_groups.append((group_start, group_len))

        # ── Entity Detail rows: (entity, sc_abbrev, rule, last_eval, entity_group_idx)
        ed_rows = []
        ed_ent_groups = []  # list of (start_idx, length) for entity merge
        for sc in all_scorecards:
            rules = sq.get(sc, {})
            if not rules:
                continue
            abbr = _abbrev(sc)
            for rule, rd in sorted(rules.items()):
                for ent in sorted(rd["entities"]):
                    ed_rows.append((ent, abbr, rule, rd["last_eval"].get(ent, "")))

        # Build entity merge groups for ED (consecutive rows with same entity)
        i = 0
        while i < len(ed_rows):
            ent = ed_rows[i][0]
            j = i + 1
            while j < len(ed_rows) and ed_rows[j][0] == ent:
                j += 1
            ed_ent_groups.append((i, j - i))
            i = j

        # ── Write all rows ────────────────────────────────────────────────────
        ROW0 = 7  # first data row
        max_rows = max(len(sc_rows), len(rb_rows), len(ed_rows), 1)

        # ABAB alternates by individual row index (0=A, 1=B, 2=A, …).
        # Merged columns (RB col H scorecard, ED col N entity) use FILL_ALT
        # consistently — no alternating on those columns.
        def _row_fill(row_idx):
            return FILL_ALT if row_idx % 2 == 0 else FILL_WHITE

        for i in range(max_rows):
            row  = ROW0 + i
            fill = _row_fill(i)          # ABAB by row for non-merged columns
            ws2.row_dimensions[row].height = 18.0

            # ── Scorecard Summary (no merged columns → ABAB on all cols) ──
            if i < len(sc_rows):
                sc_r = sc_rows[i]
                is_total = sc_r[6]
                f    = F_TOTAL_DARK if is_total else F_DARK_10
                sfill = FILL_TOTAL  if is_total else fill
                _apply(ws2.cell(row, SC_COL), value=sc_r[0], font=f, fill=sfill,
                       align=Alignment(horizontal="left", vertical="center"))
                for ci, v in enumerate(sc_r[1:6], SC_COL + 1):
                    _apply(ws2.cell(row, ci), value=v, font=f, fill=sfill,
                           align=Alignment(horizontal="center", vertical="center"))

            # ── Rule Breakdown ─────────────────────────────────────────────
            # Col H (scorecard) is merged → FILL_ALT always; cols I-L → ABAB
            if i < len(rb_rows):
                rb_r     = rb_rows[i]
                en_count = rb_r[5]
                row_h    = max(18.0, 15.0 * en_count)
                ws2.row_dimensions[row].height = row_h
                # Merged column: consistent fill
                _apply(ws2.cell(row, RB_COL),   value=rb_r[0], font=F_DARK_10, fill=FILL_ALT,
                       align=Alignment(horizontal="left", vertical="top", wrap_text=True))
                # Non-merged columns: ABAB
                _apply(ws2.cell(row, RB_COL+1), value=rb_r[1], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="left", vertical="top", wrap_text=True))
                _apply(ws2.cell(row, RB_COL+2), value=rb_r[2], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="center", vertical="top"))
                _apply(ws2.cell(row, RB_COL+3), value=rb_r[3], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="center", vertical="top"))
                _apply(ws2.cell(row, RB_COL+4), value=rb_r[4], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="left", vertical="top", wrap_text=True))

            # ── Entity Detail ──────────────────────────────────────────────
            # Col N (entity) is merged → FILL_ALT always; cols O-Q → ABAB
            if i < len(ed_rows):
                ed_r = ed_rows[i]
                # Merged column: consistent fill
                _apply(ws2.cell(row, ED_COL),   value=ed_r[0], font=F_DARK_10, fill=FILL_ALT,
                       align=Alignment(horizontal="left", vertical="center"))
                # Non-merged columns: ABAB
                _apply(ws2.cell(row, ED_COL+1), value=ed_r[1], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="left", vertical="center"))
                _apply(ws2.cell(row, ED_COL+2), value=ed_r[2], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
                _apply(ws2.cell(row, ED_COL+3), value=ed_r[3], font=F_DARK_10, fill=fill,
                       align=Alignment(horizontal="left", vertical="center"))

        # ── Merge scorecard cells in Rule Breakdown ───────────────────────────
        for g_start, g_len in rb_sc_groups:
            if g_len > 1:
                r1 = ROW0 + g_start
                r2 = ROW0 + g_start + g_len - 1
                ws2.merge_cells(start_row=r1, start_column=RB_COL,
                                end_row=r2,   end_column=RB_COL)
                ws2.cell(r1, RB_COL).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)

        # ── Merge entity cells in Entity Detail ───────────────────────────────
        for g_start, g_len in ed_ent_groups:
            if g_len > 1:
                r1 = ROW0 + g_start
                r2 = ROW0 + g_start + g_len - 1
                ws2.merge_cells(start_row=r1, start_column=ED_COL,
                                end_row=r2,   end_column=ED_COL)
                ws2.cell(r1, ED_COL).alignment = Alignment(
                    horizontal="left", vertical="center")

    # Save to buffer first, then patch chart XML, then write final file
    buf = io.BytesIO()
    wb.save(buf)
    patched = _patch_chart_xml(
        buf.getvalue(),
        cat_range=_chart_cat_range,
        title_ents=_chart_title_ents,
        title_rows=_chart_title_rows,
    )
    with open(output_xlsx, "wb") as fh:
        fh.write(patched)

    total_rows = sum(
        rd["rows"] for sq in data.values() for sc in sq.values() for rd in sc.values()
    )
    return {"squads": len(squads), "rows": total_rows}


def _patch_chart_xml(xlsx_bytes: bytes, cat_range: str, title_ents: str, title_rows: str) -> bytes:
    """
    Post-process chart XML to match 2026-06-20 reference:
    - Replace <cat><numRef> with <cat><strRef> so squad names appear on the axis
    - Replace inline <tx><v> series titles with <tx><strRef> pointing to header cells
    - Fix valAx axPos from 'l' to 'b' (bottom axis for values)
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zin, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if re.match(r"xl/charts/chart\d+\.xml", item.filename):
                xml = data.decode("utf-8")
                # categories: numRef → strRef
                xml = xml.replace(
                    f"<cat><numRef><f>{cat_range}</f></numRef></cat>",
                    f"<cat><strRef><f>{cat_range}</f></strRef></cat>",
                )
                # series 0 title: inline v → strRef
                xml = xml.replace(
                    "<tx><v>Entities</v></tx>",
                    f"<tx><strRef><f>{title_ents}</f></strRef></tx>",
                )
                # series 1 title: inline v → strRef
                xml = xml.replace(
                    "<tx><v>Rows</v></tx>",
                    f"<tx><strRef><f>{title_rows}</f></strRef></tx>",
                )
                # catAx: add <delete val="0"/> and <tickLblPos val="nextTo"/> which
                # openpyxl omits — without <delete val="0"> Excel hides the axis labels
                xml = xml.replace(
                    '<catAx><axId val="10"/><scaling><orientation val="minMax"/></scaling><axPos val="l"/>',
                    '<catAx><axId val="10"/><scaling><orientation val="minMax"/></scaling>'
                    '<delete val="0"/><axPos val="l"/><majorTickMark val="none"/>'
                    '<minorTickMark val="none"/><tickLblPos val="nextTo"/>',
                )
                # Remove the duplicate majorTickMark/minorTickMark that openpyxl adds after axPos
                xml = xml.replace(
                    '<delete val="0"/><axPos val="l"/><majorTickMark val="none"/>'
                    '<minorTickMark val="none"/><tickLblPos val="nextTo"/>'
                    '<majorTickMark val="none"/><minorTickMark val="none"/>',
                    '<delete val="0"/><axPos val="l"/><majorTickMark val="none"/>'
                    '<minorTickMark val="none"/><tickLblPos val="nextTo"/>',
                )
                # valAx axPos: catAx has axPos="l" (left), valAx must be "b" (bottom)
                # Replace the second occurrence only
                parts = xml.split('<axPos val="l"/>')
                if len(parts) == 3:
                    xml = parts[0] + '<axPos val="l"/>' + parts[1] + '<axPos val="b"/>' + parts[2]
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    return buf.getvalue()


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: build_report.py <sorted_csv> <output_xlsx>")
        sys.exit(1)
    result = build_report(sys.argv[1], sys.argv[2])
    print(f"Excel report written: {result['squads']} squads, {result['rows']} rows → {sys.argv[2]}")
